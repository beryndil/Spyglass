#!/usr/bin/env python3
"""
Extract version/edition data from Java & Bedrock xlsx databases,
cross-reference against blocks.json and items.json, and produce version_tags.json.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ASSETS = Path(__file__).resolve().parent.parent / "app/src/main/assets/minecraft"
JAVA_XLSX = Path.home() / "Downloads/minecraft_database.xlsx"
BEDROCK_XLSX = Path.home() / "Downloads/minecraft_bedrock_database.xlsx"
OUTPUT = ASSETS / "version_tags.json"


def load_game_ids():
    """Load all block and item IDs from our JSON data files."""
    blocks = {b["id"] for b in json.loads((ASSETS / "blocks.json").read_text())}
    items = {i["id"] for i in json.loads((ASSETS / "items.json").read_text())}
    return blocks, items


def namespace_to_id(ns: str) -> str:
    """Convert 'minecraft:stone_bricks' -> 'stone_bricks'."""
    if ns and ":" in ns:
        return ns.split(":", 1)[1]
    return ns or ""


def name_to_id(name: str) -> str:
    """Convert display name to snake_case ID as fallback."""
    s = name.strip().lower()
    s = re.sub(r"[''']", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s


def normalize_java_version(raw: str) -> str:
    """Normalize Java Edition version strings to our simplified format.
    Pre-release (Alpha, Beta, Classic, Indev, Infdev, rd-*) -> '1.0'
    '1.0.0' -> '1.0', '1.3.1' -> '1.3', '1.4.2' -> '1.4', etc.
    Sub-versions like '1.21.2' stay as '1.21.2' if significant.
    """
    if not raw:
        return ""
    raw = raw.strip()
    # Pre-release versions
    for prefix in ("rd-", "Classic", "Indev", "Infdev", "Alpha", "Beta"):
        if raw.startswith(prefix):
            return "1.0"

    # Handle "26.1" style (likely future version typo, keep as-is)
    # Handle standard versions
    m = re.match(r"^(\d+\.\d+)(?:\.(\d+))?$", raw)
    if m:
        major_minor = m.group(1)
        patch = m.group(2)
        # For older versions, drop patch (1.3.1 -> 1.3, 1.4.2 -> 1.4)
        # For 1.20+ keep significant sub-versions (1.20.5, 1.21.2, etc.)
        if patch and patch != "0":
            parts = major_minor.split(".")
            minor = int(parts[1]) if len(parts) > 1 else 0
            if minor >= 20:
                return f"{major_minor}.{patch}"
            # Older versions: 1.2.1->1.2, 1.3.1->1.3, 1.4.2->1.4, etc.
            return major_minor
        return major_minor

    # Handle versions like "1.0.0" -> "1.0"
    m2 = re.match(r"^(\d+)\.(\d+)\.0$", raw)
    if m2:
        return f"{m2.group(1)}.{m2.group(2)}"

    return raw


def normalize_bedrock_version(raw: str) -> str:
    """Normalize Bedrock Edition version strings.
    'PE Alpha 0.X.Y' -> '0.X'
    'PE 1.0.0' / 'BE 1.0.0' -> '1.0'
    'BE 1.16.0' -> '1.16'
    """
    if not raw:
        return ""
    raw = raw.strip()

    # PE Alpha versions -> simplified
    m = re.match(r"PE Alpha (\d+)\.(\d+)", raw)
    if m:
        return f"{m.group(1)}.{m.group(2)}"

    # PE/BE prefix
    for prefix in ("PE ", "BE "):
        if raw.startswith(prefix):
            raw = raw[len(prefix):]
            break

    # Standard version normalization
    m = re.match(r"^(\d+\.\d+)(?:\.(\d+))?$", raw)
    if m:
        major_minor = m.group(1)
        patch = m.group(2)
        if patch and patch != "0":
            parts = major_minor.split(".")
            minor = int(parts[1]) if len(parts) > 1 else 0
            if minor >= 20:
                return f"{major_minor}.{patch}"
            return major_minor
        return major_minor

    return raw


# Special name -> ID mappings for tricky cases
SPECIAL_MAPPINGS = {
    "redstone dust": "redstone",
    "nether wart": "nether_wart",
    "nether warts": "nether_wart",
    "jack o'lantern": "jack_o_lantern",
    "jack o' lantern": "jack_o_lantern",
    "turtle shell (scute)": "turtle_scute",
    "turtle scute": "turtle_scute",
    "scute": "turtle_scute",
    "tnt": "tnt",
}


def read_java_xlsx():
    """Read Java Edition xlsx and return list of dicts."""
    wb = openpyxl.load_workbook(JAVA_XLSX, read_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        name = str(row[1]).strip()
        entity_type = str(row[2] or "").strip().lower()
        namespace_id = str(row[10] or "").strip() if row[10] else ""
        version_introduced = str(row[4] or "").strip() if row[4] else ""
        mechanics_changed = str(row[7] or "").strip() if row[7] else ""
        mechanics_notes = str(row[8] or "").strip() if row[8] else ""
        status = str(row[9] or "").strip() if row[9] else ""

        # Determine entity type
        if entity_type == "block":
            etype = "block"
        elif entity_type == "item":
            etype = "item"
        else:
            etype = "block"  # default for ambiguous

        # Determine ID
        eid = namespace_to_id(namespace_id)
        if not eid:
            low = name.lower()
            eid = SPECIAL_MAPPINGS.get(low, name_to_id(name))

        entries.append({
            "name": name,
            "entityType": etype,
            "entityId": eid,
            "addedInJava": normalize_java_version(version_introduced),
            "mechanicsChangedInJava": normalize_java_version(mechanics_changed),
            "mechanicsChangeNotes": mechanics_notes if mechanics_notes and mechanics_notes.lower() != "none" else "",
            "status": status,
        })
    wb.close()
    return entries


def read_bedrock_xlsx():
    """Read Bedrock Edition xlsx and return dict keyed by namespace ID."""
    wb = openpyxl.load_workbook(BEDROCK_XLSX, read_only=True)
    ws = wb.active
    entries = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        name = str(row[1]).strip()
        namespace_id = str(row[12] or "").strip() if row[12] else ""
        be_version = str(row[4] or "").strip() if row[4] else ""
        mechanics_changed = str(row[8] or "").strip() if row[8] else ""
        mechanics_notes = str(row[9] or "").strip() if row[9] else ""
        be_exclusive = str(row[11] or "").strip().lower() if row[11] else ""

        eid = namespace_to_id(namespace_id)
        if not eid:
            low = name.lower()
            eid = SPECIAL_MAPPINGS.get(low, name_to_id(name))

        entries[eid] = {
            "addedInBedrock": normalize_bedrock_version(be_version),
            "mechanicsChangedInBedrock": normalize_bedrock_version(mechanics_changed),
            "mechanicsChangeNotes": mechanics_notes if mechanics_notes and mechanics_notes.lower() != "none" else "",
            "beExclusive": be_exclusive in ("yes", "true", "1"),
        }
    wb.close()
    return entries


def main():
    blocks, items = load_game_ids()
    all_ids = blocks | items
    print(f"Game data: {len(blocks)} blocks, {len(items)} items, {len(all_ids)} total")

    java_entries = read_java_xlsx()
    bedrock_data = read_bedrock_xlsx()
    print(f"Java xlsx: {len(java_entries)} entries")
    print(f"Bedrock xlsx: {len(bedrock_data)} entries")

    # Build output, keyed by (entityType, entityId) to deduplicate
    output = {}
    matched_ids = set()
    unmatched_java = []

    for entry in java_entries:
        eid = entry["entityId"]

        # Determine actual entity type based on our data
        if eid in blocks and eid in items:
            etype = entry["entityType"]  # use xlsx type hint
        elif eid in blocks:
            etype = "block"
        elif eid in items:
            etype = "item"
        else:
            unmatched_java.append(f"  {entry['name']} -> {eid} (xlsx type: {entry['entityType']})")
            continue

        key = (etype, eid)
        if key in output:
            continue  # skip duplicates

        matched_ids.add(eid)

        be = bedrock_data.get(eid, {})
        added_bedrock = be.get("addedInBedrock", "")
        be_exclusive = be.get("beExclusive", False)
        mechanics_bedrock = be.get("mechanicsChangedInBedrock", "")

        # Merge mechanics notes from both
        java_notes = entry.get("mechanicsChangeNotes", "")
        be_notes = be.get("mechanicsChangeNotes", "")
        # Prefer Java notes; append Bedrock notes if different
        if be_notes and be_notes != java_notes:
            notes = f"{java_notes}; BE: {be_notes}" if java_notes else be_notes
        else:
            notes = java_notes

        java_only = eid not in bedrock_data and not be_exclusive
        bedrock_only = be_exclusive

        tag = {
            "entityType": etype,
            "entityId": eid,
        }

        # Only include non-default fields to keep JSON compact
        added_java = entry.get("addedInJava", "")
        if added_java and added_java != "1.0":
            tag["addedInJava"] = added_java
        if added_bedrock:
            tag["addedInBedrock"] = added_bedrock
        if java_only:
            tag["javaOnly"] = True
        if bedrock_only:
            tag["bedrockOnly"] = True
        mech_java = entry.get("mechanicsChangedInJava", "")
        if mech_java:
            tag["mechanicsChangedInJava"] = mech_java
        if mechanics_bedrock:
            tag["mechanicsChangedInBedrock"] = mechanics_bedrock
        if notes:
            tag["mechanicsChangeNotes"] = notes

        output[key] = tag

    # Add Bedrock-exclusive items not in Java xlsx
    for eid, be in bedrock_data.items():
        if eid in matched_ids:
            continue
        if be.get("beExclusive"):
            if eid in blocks:
                etype = "block"
            elif eid in items:
                etype = "item"
            else:
                continue
            key = (etype, eid)
            if key not in output:
                tag = {"entityType": etype, "entityId": eid, "bedrockOnly": True}
                if be.get("addedInBedrock"):
                    tag["addedInBedrock"] = be["addedInBedrock"]
                output[key] = tag
                matched_ids.add(eid)

    # Find items in our data that weren't in either xlsx (they're "since 1.0" in both)
    unmatched_game = all_ids - matched_ids

    # Sort output by entityType then entityId
    result = sorted(output.values(), key=lambda x: (x["entityType"], x["entityId"]))

    # Write output
    OUTPUT.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n")

    print(f"\nOutput: {len(result)} entries written to {OUTPUT}")
    print(f"Matched: {len(matched_ids)} of {len(all_ids)} game IDs")
    print(f"Unmatched from xlsx: {len(unmatched_java)}")
    if unmatched_java:
        print("First 20 unmatched from xlsx:")
        for line in unmatched_java[:20]:
            print(line)
    print(f"Game IDs with no xlsx entry (assumed since 1.0): {len(unmatched_game)}")
    if unmatched_game:
        print("First 20:", sorted(unmatched_game)[:20])


if __name__ == "__main__":
    main()
